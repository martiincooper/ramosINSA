"""Write the convalidation workbook (problem statement section 8).

The workbook now includes a sixth sheet, ``Convalidation Form``, that reproduces
the official USM equivalence table (destination university course <-> USM
equivalent) exactly as required by the printed convalidation form.
"""
from __future__ import annotations

import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import config
from .config import RULES
from .models import CandidateMatch, INSACourse, Recommendation, USMCourse
from .study_plan import StudyPlanRow

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Validation statuses that represent an actual proposed convalidation (i.e. one
# that should appear on the official equivalence form).
_ACCEPTED_STATUSES = {
    RULES.status_valid,
    RULES.status_valid_combination,
    RULES.status_borderline,
    RULES.status_needs_work,
}


def _write_sheet(ws, headers: List[str], rows: List[list]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    _autosize(ws, len(headers))


def _autosize(ws, n_cols: int, max_width: int = 60) -> None:
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        longest = 0
        for cell in ws[letter]:
            value = "" if cell.value is None else str(cell.value)
            longest = max(longest, min(len(value), max_width))
        ws.column_dimensions[letter].width = max(12, longest + 2)
        for cell in ws[letter]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_workbook(
    usm_courses: List[USMCourse],
    insa_courses: List[INSACourse],
    candidates: List[CandidateMatch],
    recommendations: List[Recommendation],
    study_plan: List[StudyPlanRow],
    output_path: str = None,
) -> str:
    output_path = output_path or config.EXCEL_OUTPUT
    wb = Workbook()

    # Sheet 1 - USM Courses
    ws = wb.active
    ws.title = "USM Courses"
    _write_sheet(
        ws,
        ["code", "name", "SCT credits", "department", "description", "key topics"],
        [
            [c.code, c.title, c.sct_credits, c.department, c.description, c.key_topics()]
            for c in usm_courses
        ],
    )

    # Sheet 2 - INSA Courses
    ws = wb.create_sheet("INSA Courses")
    _write_sheet(
        ws,
        ["code", "name", "ECTS", "year", "semester", "department", "key topics"],
        [
            [c.code, c.title, c.ects, c.year, c.semester, c.department, c.key_topics()]
            for c in insa_courses
        ],
    )

    # Sheet 3 - Candidate Matches
    ws = wb.create_sheet("Candidate Matches")
    _write_sheet(
        ws,
        ["USM course", "INSA course", "similarity score", "ECTS", "notes"],
        [
            [
                f"{c.usm_code}",
                f"{c.insa_code} - {c.insa_title}",
                round(c.similarity, 4),
                c.ects,
                c.notes or c.department,
            ]
            for c in candidates
        ],
    )

    # Sheet 4 - Recommended Convalidations
    ws = wb.create_sheet("Recommended Convalidations")
    _write_sheet(
        ws,
        [
            "USM course",
            "recommended INSA course(s)",
            "combined ECTS",
            "estimated equivalence %",
            "validation status",
            "justification",
        ],
        [
            [
                f"{r.usm_code} - {r.usm_title}",
                _format_insa_list(r),
                r.combined_ects,
                round(r.equivalence * 100, 1),
                r.status,
                r.justification,
            ]
            for r in recommendations
        ],
    )

    # Sheet 5 - Final Proposed Study Plan
    ws = wb.create_sheet("Final Proposed Study Plan")
    _write_sheet(
        ws,
        [
            "semester",
            "INSA courses",
            "total ECTS",
            "departments involved",
            "target USM convalidations",
            "notes",
        ],
        [
            [
                row.semester,
                "\n".join(row.insa_courses),
                row.total_ects,
                "\n".join(row.departments),
                ", ".join(row.target_usm),
                "\n".join(row.warnings),
            ]
            for row in study_plan
        ],
    )

    # Sheet 6 - Official USM convalidation form (matches the printed template)
    ws = wb.create_sheet("Convalidation Form")
    _write_convalidation_form(ws, usm_courses, insa_courses, recommendations)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# --------------------------------------------------------------------------- #
# Official USM convalidation form (destination university <-> USM equivalent)
# --------------------------------------------------------------------------- #
_FORM_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_CHECKED = "\u2611"  # ballot box with check
_UNCHECKED = "\u2610"  # ballot box


def _insa_level_cell(course: INSACourse) -> str:
    """Render the NIVEL ASIGNATURA checkbox column (Pregrado / Master).

    INSA Lyon 4th- and 5th-year course sheets correspond to Master level; the
    earlier years are Pregrado.
    """
    is_master = str(course.year).strip() in RULES.master_level_years
    pregrado = _CHECKED if not is_master else _UNCHECKED
    master = _CHECKED if is_master else _UNCHECKED
    return f"{pregrado} Pregrado\n{master} Master"


def _write_convalidation_form(
    ws,
    usm_courses: List[USMCourse],
    insa_courses: List[INSACourse],
    recommendations: List[Recommendation],
) -> None:
    """Write the official equivalence form matching the USM convalidation table.

    Layout (two tier header):

        | ASIGNATURA EN UNIVERSIDAD DE DESTINO        | ASIGNATURA EQUIVALENTE USM |
        | SIGLA | NOMBRE | IDIOMA | NIVEL | CRÉDITOS   | SIGLA | NOMBRE | CRÉDITOS SCT |

    The "universidad de destino" is INSA Lyon. Each row is one INSA (destination)
    course; when a USM course is convalidated by a combination of several INSA
    courses they appear on consecutive rows and the USM equivalent is shown only
    on the first of those rows.
    """
    insa_by_code = {c.code: c for c in insa_courses}
    usm_by_code = {c.code: c for c in usm_courses}

    # Tier 1 (grouped) header.
    ws.merge_cells("A1:E1")
    ws.merge_cells("F1:H1")
    ws["A1"] = "ASIGNATURA EN UNIVERSIDAD DE DESTINO"
    ws["F1"] = "ASIGNATURA EQUIVALENTE USM"

    # Tier 2 (column) header.
    sub_headers = [
        "SIGLA",
        "NOMBRE ASIGNATURA",
        "IDIOMA INSTRUCCIÓN",
        "NIVEL ASIGNATURA",
        "CRÉDITOS",
        "SIGLA",
        "NOMBRE ASIGNATURA",
        "CRÉDITOS SCT",
    ]
    ws.append(sub_headers)

    for rec in recommendations:
        if not rec.insa_codes or rec.status not in _ACCEPTED_STATUSES:
            continue
        usm = usm_by_code.get(rec.usm_code)
        for position, code in enumerate(rec.insa_codes):
            insa = insa_by_code.get(code)
            insa_title = insa.title if insa else (
                rec.insa_titles[position] if position < len(rec.insa_titles) else code
            )
            row = [
                code,
                insa_title,
                insa.teaching_language if insa else "",
                _insa_level_cell(insa) if insa else "",
                insa.ects if insa else "",
            ]
            if position == 0:
                row += [
                    rec.usm_code,
                    usm.title if usm else rec.usm_title,
                    usm.sct_credits if usm else "",
                ]
            else:
                # Continuation row of a multi-course combination: the USM
                # equivalent is already shown on the first row.
                row += ["", "", ""]
            ws.append(row)

    _autosize(ws, len(sub_headers))

    # Re-apply the form styling (autosize resets alignment for every cell).
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(sub_headers)):
        for cell in row:
            cell.border = _FORM_BORDER
            if cell.row <= 2:
                cell.fill = _HEADER_FILL
                cell.font = _HEADER_FONT
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A3"


def _format_insa_list(rec: Recommendation) -> str:
    if not rec.insa_codes:
        return "-"
    return "\n".join(
        f"{code} - {title}" for code, title in zip(rec.insa_codes, rec.insa_titles)
    )
